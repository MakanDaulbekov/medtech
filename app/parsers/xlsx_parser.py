"""
XLSX Parser
- Iterates all sheets
- Detects header row automatically (not necessarily row 0)
- Maps service name and price columns heuristically
"""
import re
import logging
from typing import List, Tuple, Optional
from decimal import Decimal, InvalidOperation

import openpyxl

logger = logging.getLogger(__name__)

MAX_HEADER_SCAN_ROWS = 15   # scan first N rows looking for headers


def _clean_price(raw) -> Optional[Decimal]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        val = Decimal(str(raw))
        return val if val > 0 else None
    cleaned = re.sub(r"[^\d.,]", "", str(raw)).replace(",", ".")
    try:
        val = Decimal(cleaned)
        return val if val > 0 else None
    except InvalidOperation:
        return None


def _score_header_row(row_values: List) -> int:
    """Score a row as candidate header. Higher = more likely a header."""
    score = 0
    keywords = [
        "услуга", "наименование", "название", "код", "цена",
        "стоимость", "резидент", "нерезидент", "тариф"
    ]
    for v in row_values:
        if isinstance(v, str):
            vl = v.lower()
            for kw in keywords:
                if kw in vl:
                    score += 1
    return score


def _is_price_col(header: str) -> str | None:
    h = str(header).lower()
    if "нерезидент" in h or "non" in h:
        return "nonresident"
    if any(k in h for k in ["цена", "стоимость", "резидент", "тариф", "сумма", "kzt", "тг"]):
        return "resident"
    return None


def _is_service_col(header: str) -> bool:
    h = str(header).lower()
    return any(k in h for k in ["услуга", "наименование", "название", "описание", "процедура"])


def parse_xlsx(file_path: str) -> Tuple[List[dict], str]:
    """Returns (items, raw_text summary)."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_items: List[dict] = []
    raw_parts: List[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # ── Find header row ───────────────────────────────────────────────────
        best_header_idx = 0
        best_score = 0
        for i, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
            score = _score_header_row([v for v in row if v is not None])
            if score > best_score:
                best_score = score
                best_header_idx = i

        headers = [str(v).strip() if v is not None else f"col_{j}"
                   for j, v in enumerate(rows[best_header_idx])]

        # ── Map columns ────────────────────────────────────────────────────────
        svc_col = None
        code_col = None
        res_col = None
        nonres_col = None

        for j, h in enumerate(headers):
            if _is_service_col(h) and svc_col is None:
                svc_col = j
            if "код" in h.lower() and code_col is None:
                code_col = j
            kind = _is_price_col(h)
            if kind == "resident" and res_col is None:
                res_col = j
            elif kind == "nonresident" and nonres_col is None:
                nonres_col = j

        if svc_col is None:
            # fallback: longest string column
            svc_col = 0

        # ── Extract rows ───────────────────────────────────────────────────────
        sheet_items = 0
        for row in rows[best_header_idx + 1:]:
            if not row or all(v is None for v in row):
                continue

            name = str(row[svc_col]).strip() if svc_col < len(row) and row[svc_col] else ""
            if not name or name.lower() in ("none", "", "nan") or len(name) < 3:
                continue

            raw_parts.append(name)
            code = str(row[code_col]).strip() if code_col is not None and code_col < len(row) and row[code_col] else None
            price_r = _clean_price(row[res_col]) if res_col is not None and res_col < len(row) else None
            price_nr = _clean_price(row[nonres_col]) if nonres_col is not None and nonres_col < len(row) else None

            # Last resort price scan
            if price_r is None:
                for j, cell in enumerate(row):
                    if j == svc_col:
                        continue
                    candidate = _clean_price(cell)
                    if candidate and Decimal("1") < candidate < Decimal("10000000"):
                        price_r = candidate
                        break

            all_items.append({
                "service_name_raw": name,
                "service_code_source": code,
                "price_resident_kzt": price_r,
                "price_nonresident_kzt": price_nr,
                "currency_original": "KZT",
            })
            sheet_items += 1

        logger.info(f"Sheet '{sheet_name}': extracted {sheet_items} items")

    wb.close()
    return all_items, "\n".join(raw_parts[:200])
